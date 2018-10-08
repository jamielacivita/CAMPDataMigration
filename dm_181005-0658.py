from openpyxl import Workbook, load_workbook
import datetime
import sys

filename = sys.argv[1]  #Filename is the first argument on the commandline.

#load workbook
#todo : add code so that you can read filename from argunments.
#wb = load_workbook('432 All Approved Access and Special Data Flags 2018-08-16.xlsx')
wb = load_workbook(filename)

#activate sheet
ws = wb.active

#create new sheet
#wsNew = wb.create_sheet(title="Entity Profiles From 432")

scanrow = 1 # we are scaning entries in row 1 only.
if ws.cell(row=scanrow, column=1).value == "EntityID":
    print("Column A is correct")
else:
    print("Column A is incorrect")







data=[('Entity Name', 'Entity Unique ID', 'Legacy 432 Entity ID', 'External Entity ID', 'Alias', 'Sourcing Company',
'Entity Country', 'Entity Risk Rating', 'Competitor', 'Subject to Export Compliance Laws',
'Contractual or Local Law Restrictions', 'High Risk Conuntry', 'Date of Last Review', 'Entity Info Type: IP Access',
'Conditions', 'Entity Info Type:  SPD Access', 'Conditions', 'Entity Info Type:  TD Access', 'Conditions',
'Data Info Classification', 'Access without a Chevron ID:',   'Access without a Chevron ID:  Additions',
'Access without a Chevron ID:  Exclusions', 'Access with a Chevron ID:', 'Access with a Chevron ID:  Additions',
'Access with a Chevron ID:  Exclusions', 'Email Access:', 'Email Access:  Additions', 'Email Access:  Exclusions',
'Shared Drive:', 'Shared Drive:  Additions', 'Shared Drive:  Exclusions', 'Intranet:', 'Intranet:  Additions',
'Intranet:  Exclusions')]

#Adding titles to columns
for row in data:
    wsNew.append(row)

# get max row count
max_row = ws.max_row

# get max column count
max_column = ws.max_column

# Function to determine if the new sheet contains the target
# Returns boolean
def contains(target):
    retval = False
    for a in range (1,wsNew.max_row+1):
        if target == wsNew.cell(row=a, column=1).value:
            retval = True
    return retval

def findRow(target):
    retval = 0
    for a in range (1,wsNew.max_row+1):
        if target == wsNew.cell(row=a, column=1).value:
            retval = a
    return retval

rowCount = 2
count = 0
# iterate over all cells
# iterate over all rows
print("Beginning iteraiton over all cells and rows\n")
for i in range(2,ws.max_row+1):
#for i in range(2,2000):
    count += 1
    if count %100 == 0:
        print(count)

    if ws.cell(row=i, column=3).value == "LEGACY" or ws.cell(row=i, column=3).value == "Legacy" or ws.cell(row=i, column=3).value == None:
        ws.cell(row=i, column=3).value = ws.cell(row=i, column=15).value
        print("\tLegacy or space found at row: " + str(i))
    #If new entry, add to sheet
    if contains(ws.cell(row=i,column=2).value) == False:
        #Entity Name
        wsNew.cell(row=rowCount, column=1).value = ws.cell(row=i,column=2).value

        #Legacy 432 Entity ID
        wsNew.cell(row=rowCount, column=3).value = ws.cell(row=i,column=1).value

        #External Entity ID
        wsNew.cell(row=rowCount, column=4).value = ws.cell(row=i,column=8).value

        #Alias
        wsNew.cell(row=rowCount, column=5).value = ws.cell(row=i,column=4).value

        #Sourcing Company
        wsNew.cell(row=rowCount, column=6).value = ws.cell(row=i,column=6).value

        #Entity Conuntry
        wsNew.cell(row=rowCount, column=7).value = ws.cell(row=i,column=3).value

        #Date of last review
        wsNew.cell(row=rowCount, column=13).value = ws.cell(row=i,column=20).value

        #Entity info type: IP Access
        wsNew.cell(row=rowCount, column=14).value = ws.cell(row=i,column=11).value

        #Entity info type: SPD Access
        wsNew.cell(row=rowCount, column=16).value = ws.cell(row=i,column=10).value

        #Entity info type: TD Access
        wsNew.cell(row=rowCount, column=18).value = ws.cell(row=i,column=9).value

        #Entity Risk Rating
        if ws.cell(row=i,column=9).value == "Yes" or ws.cell(row=i,column=10).value == "Yes" or ws.cell(row=i,column=11).value == "Yes":
            wsNew.cell(row=rowCount,column=8).value = 'high'

        #Shared Drive
        if ws.cell(row=i,column=13).value == "Shared Drive":
            wsNew.cell(row=rowCount,column=30).value = 'yes'
            wsNew.cell(row=rowCount,column=24).value = 'yes'

        #Access without Chevron ID
        if ws.cell(row=i,column=14).value == "Application Gateway":
            wsNew.cell(row=rowCount,column=21).value = 'yes'

        #Contractor Basic Access
        if ws.cell(row=i,column=14).value == "Contractor Basic Access":
            wsNew.cell(row=rowCount,column=33).value = 'basic'
            wsNew.cell(row=rowCount,column=24).value = 'yes'

        #Contractor Full Access
        if ws.cell(row=i,column=14).value == "Contractor Full Access (Selected Contractor)":
            wsNew.cell(row=rowCount,column=33).value = 'full'
            wsNew.cell(row=rowCount,column=24).value = 'yes'

        #Email
        if ws.cell(row=i,column=14).value == "Chevron E-Mail(Outlook)":
            wsNew.cell(row=rowCount,column=24).value = 'yes'
            wsNew.cell(row=rowCount,column=27).value = 'yes'

        #Access with Chevron ID
        if ws.cell(row=i,column=14).value == "Chevron Intranet" or ws.cell(row=i,column=14).value == "CT Account" or ws.cell(row=i,column=14).value == "GIL Machine/Laptop" or ws.cell(row=i,column=14).value == "LMS (Learning Management System)" or ws.cell(row=i,column=14).value == "CAT (Compliance Activity Tracker)" or ws.cell(row=i,column=14).value == "P Drive":
            wsNew.cell(row=rowCount,column=24).value = 'yes'

        # P Drive
        if ws.cell(row=i,column=14).value == "P Drive":
            wsNew.cell(row=rowCount,column=24).value = 'yes'
            wsNew.cell(row=rowCount,column=30).value = 'yes'

        #Increment New Sheet Row
        rowCount += 1

    #If entry exists, edit existing entry
    else:
        editRow = findRow(ws.cell(row=i,column=2).value)

        #Creating new entry when country is different
        if ws.cell(row=i, column=3).value != wsNew.cell(row=editRow, column=7).value:
            #Entity Name
            wsNew.cell(row=rowCount, column=1).value = ws.cell(row=i,column=2).value

            #Legacy 432 Entity ID
            wsNew.cell(row=rowCount, column=3).value = ws.cell(row=i,column=1).value

            #External Entity ID
            wsNew.cell(row=rowCount, column=4).value = ws.cell(row=i,column=8).value

            #Alias
            wsNew.cell(row=rowCount, column=5).value = ws.cell(row=i,column=4).value

            #Sourcing Company
            wsNew.cell(row=rowCount, column=6).value = ws.cell(row=i,column=6).value

            #Entity Conuntry
            wsNew.cell(row=rowCount, column=7).value = ws.cell(row=i,column=3).value

            #Date of last review
            wsNew.cell(row=rowCount, column=13).value = ws.cell(row=i,column=20).value

            #Entity info type: IP Access
            wsNew.cell(row=rowCount, column=14).value = ws.cell(row=i,column=11).value

            #Entity info type: SPD Access
            wsNew.cell(row=rowCount, column=16).value = ws.cell(row=i,column=10).value

            #Entity info type: TD Access
            wsNew.cell(row=rowCount, column=18).value = ws.cell(row=i,column=9).value

            #Entity Risk Rating
            if ws.cell(row=i,column=9).value == "Yes" or ws.cell(row=i,column=10).value == "Yes" or ws.cell(row=i,column=11).value == "Yes":
                wsNew.cell(row=rowCount,column=8).value = 'high'

            #Shared Drive
            if ws.cell(row=i,column=13).value == "Shared Drive":
                wsNew.cell(row=rowCount,column=30).value = 'yes'
                wsNew.cell(row=rowCount,column=24).value = 'yes'

            #Access without Chevron ID
            if ws.cell(row=i,column=14).value == "Application Gateway":
                wsNew.cell(row=rowCount,column=21).value = 'yes'

            #Contractor Basic Access
            if ws.cell(row=i,column=14).value == "Contractor Basic Access":
                wsNew.cell(row=rowCount,column=33).value = 'basic'
                wsNew.cell(row=rowCount,column=24).value = 'yes'

            #Contractor Full Access
            if ws.cell(row=i,column=14).value == "Contractor Full Access (Selected Contractor)":
                wsNew.cell(row=rowCount,column=33).value = 'full'
                wsNew.cell(row=rowCount,column=24).value = 'yes'

            #Email
            if ws.cell(row=i,column=14).value == "Chevron E-Mail(Outlook)":
                wsNew.cell(row=rowCount,column=24).value = 'yes'
                wsNew.cell(row=rowCount,column=27).value = 'yes'

            #Access with Chevron ID
            if ws.cell(row=i,column=14).value == "Chevron Intranet" or ws.cell(row=i,column=14).value == "CT Account" or ws.cell(row=i,column=14).value == "GIL Machine/Laptop" or ws.cell(row=i,column=14).value == "LMS (Learning Management System)" or ws.cell(row=i,column=14).value == "CAT (Compliance Activity Tracker)" or ws.cell(row=i,column=14).value == "P Drive":
                wsNew.cell(row=rowCount,column=24).value = 'yes'

            # P Drive
            if ws.cell(row=i,column=14).value == "P Drive":
                wsNew.cell(row=rowCount,column=24).value = 'yes'
                wsNew.cell(row=rowCount,column=30).value = 'yes'

            #Increment New Sheet Row
            rowCount += 1

        #Updating Existing row in new Sheet
        else:
            #Update: Date of last review
            #if type(wsNew.cell(row=editRow, column=13).value) == datetime.datetime and type(ws.cell(row=i,column=20).value) == datetime.datetime:
            if isinstance(ws.cell(row=i,column=20).value,datetime.datetime) and isinstance(wsNew.cell(row=editRow, column=13).value,datetime.datetime):
                if ws.cell(row=i,column=20).value > wsNew.cell(row=editRow, column=13).value:
                    wsNew.cell(row=editRow, column=13).value = ws.cell(row=i,column=20).value

            #Update: Entity info type: IP Access
            wsNew.cell(row=editRow, column=14).value = ws.cell(row=i,column=11).value

            #Update: Entity info type: SPD Access
            wsNew.cell(row=editRow, column=16).value = ws.cell(row=i,column=10).value

            #Update: Entity info type: TD Access
            wsNew.cell(row=editRow, column=18).value = ws.cell(row=i,column=9).value

            #Entity Risk Rating
            if ws.cell(row=i,column=9).value == "Yes" or ws.cell(row=i,column=10).value == "Yes" or ws.cell(row=i,column=11).value == "Yes":
                wsNew.cell(row=editRow,column=8).value = 'high'

            #Shared Drive
            if ws.cell(row=i,column=13).value == "Shared Drive":
                wsNew.cell(row=editRow,column=30).value = 'yes'
                wsNew.cell(row=editRow,column=24).value = 'yes'

            #Access without Chevron ID
            if ws.cell(row=i,column=14).value == "Application Gateway":
                wsNew.cell(row=editRow,column=21).value = 'yes'

            #Contractor Basic Access
            if ws.cell(row=i,column=14).value == "Contractor Basic Access":
                wsNew.cell(row=editRow,column=33).value = 'basic'
                wsNew.cell(row=editRow,column=24).value = 'yes'

            #Contractor Full Access
            if ws.cell(row=i,column=14).value == "Contractor Full Access (Selected Contractor)":
                wsNew.cell(row=editRow,column=33).value = 'full'
                wsNew.cell(row=editRow,column=24).value = 'yes'

            #Access with Chevron ID and Email
            if ws.cell(row=i,column=14).value == "Chevron E-Mail (Outlook)":
                wsNew.cell(row=editRow,column=24).value = 'yes'
                wsNew.cell(row=editRow,column=27).value = 'yes'

            #Access with Chevron ID
            if ws.cell(row=i,column=14).value == "Chevron Intranet" or ws.cell(row=i,column=14).value == "CT Account" or ws.cell(row=i,column=14).value == "GIL Machine/Laptop" or ws.cell(row=i,column=14).value == "LMS (Learning Management System)" or ws.cell(row=i,column=14).value == "CAT (Compliance Activity Tracker)":
                wsNew.cell(row=editRow,column=24).value = 'yes'

            # P Drive
            if ws.cell(row=i,column=14).value == "P Drive":
                wsNew.cell(row=editRow,column=24).value = 'yes'
                wsNew.cell(row=editRow,column=30).value = 'yes'



wb.save('EntityProfilesFrom432.xlsx')
